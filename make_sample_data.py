"""Generates sample_data.xlsx for local testing of the full pipeline.

Layout matches what generate_pricelist.py auto-detects on the "Price List"
sheet: column A = item name, column B = description, column C = CHF price.
A row with only column A filled is a category header.

Also includes a legacy "Exchange Rate" sheet, purely for reference/backward
compatibility with older workbooks -- generate_pricelist.py never reads it.
The live-fetched rate (fx_rate.py) is the only source of truth for EUR
prices now.
"""

import openpyxl
from openpyxl.styles import Font

CATEGORIES = [
    ("Espresso Machines", [
        ("Compatto Espresso Machine", "Single-boiler, PID temperature control", 890.00),
        ("Duetto Espresso Machine", "Dual-boiler, rotary pump", 1650.00),
        ("Portable Espresso Maker", "Manual lever, travel case included", 129.00),
    ]),
    ("Grinders", [
        ("Conical Burr Grinder", "Stepless adjustment, 250g hopper", 340.00),
        ("Flat Burr Grinder", "58mm burrs, single-dose", 610.00),
        ("Hand Grinder", "Steel burrs, compact travel size", 89.00),
    ]),
    ("Filter Coffee", [
        ("Pour-Over Dripper", "Ceramic, size 02", 42.00),
        ("Batch Brewer", "1.8L capacity, glass carafe", 210.00),
        ("Cold Brew Kit", "Glass carafe with reusable filter", 58.00),
    ]),
    ("Kettles", [
        ("Gooseneck Kettle", "Variable temperature, 0.9L", 145.00),
        ("Stovetop Kettle", "Stainless steel, gooseneck spout", 68.00),
    ]),
    ("Cups & Mugs", [
        ("Espresso Cup Set", "Set of 2, porcelain", 36.00),
        ("Cappuccino Cup", "Double-walled glass", 22.00),
        ("Travel Mug", "Insulated, 350ml", 34.00),
    ]),
    ("Scales", [
        ("Brewing Scale", "0.1g precision, built-in timer", 79.00),
        ("Espresso Scale", "Compact, drip-tray sized", 95.00),
    ]),
    ("Tampers & Accessories", [
        ("Calibrated Tamper", "58mm, spring-loaded", 65.00),
        ("Distribution Tool", "Adjustable, 58mm", 45.00),
        ("Knock Box", "Stainless steel, removable bar", 38.00),
    ]),
    ("Water Filtration", [
        ("Countertop Filter", "Reduces scale and chlorine", 120.00),
        ("Replacement Cartridge", "3-month supply", 24.00),
    ]),
    ("Milk Frothing", [
        ("Milk Frothing Pitcher 350ml", "Stainless steel, pointed spout", 28.00),
        ("Milk Frothing Pitcher 600ml", "Stainless steel, pointed spout", 32.00),
        ("Handheld Milk Frother", "Battery-powered, whisk attachment", 19.00),
    ]),
    ("Coffee Beans", [
        ("House Blend 250g", "Medium roast, whole bean", 14.00),
        ("Single Origin 250g", "Light roast, seasonal", 16.50),
        ("Decaf Blend 250g", "Swiss water process", 15.00),
    ]),
    ("Tea", [
        ("Loose Leaf Sencha 100g", "Japanese green tea", 12.00),
        ("Loose Leaf Earl Grey 100g", "Bergamot-scented black tea", 11.00),
    ]),
    ("Storage", [
        ("Coffee Canister", "Airtight, 500g capacity", 32.00),
        ("Tea Tin Set", "Set of 3, airtight", 28.00),
    ]),
    ("Cleaning Supplies", [
        ("Espresso Machine Cleaner", "Backflush detergent, 430g", 18.00),
        ("Descaling Solution", "500ml, for boilers and kettles", 16.00),
        ("Group Head Brush", "Wooden handle, nylon bristles", 9.00),
    ]),
    ("Merchandise", [
        ("Branded Apron", "Cotton canvas, adjustable strap", 42.00),
        ("Branded Tote Bag", "Heavy canvas, screen-printed logo", 24.00),
        ("Branded Cap", "Adjustable, embroidered logo", 26.00),
    ]),
]


def build_price_list_sheet(workbook: openpyxl.Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Price List"
    # Category header rows are detected by bold formatting on column A (see
    # generate_pricelist.py) -- no separate "Item/Description/CHF" title row
    # is needed. Column C on the header row labels the price column itself.
    # Deliberately not labelled "CHF" here -- that string doubles as the
    # currency-column header in the PDF, so reusing it as an option label
    # would make text-based currency checks ambiguous. "Price" is neutral.
    for category_name, items in CATEGORIES:
        sheet.append([category_name, None, "Price"])
        header_cell = sheet.cell(row=sheet.max_row, column=1)
        header_cell.font = Font(bold=True, size=13)
        for name, description, chf in items:
            sheet.append([name, description, chf])
        sheet.append([None, None, None])  # blank separator between categories

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 42
    sheet.column_dimensions["C"].width = 12


def build_legacy_exchange_rate_sheet(workbook: openpyxl.Workbook) -> None:
    """Kept for reference/backward compatibility only -- generate_pricelist.py
    never reads this sheet. EUR prices are now always computed from the
    live-fetched rate at generation time (see fx_rate.py).
    """
    sheet = workbook.create_sheet("Exchange Rate")
    sheet.append(["Legacy manual exchange rate (no longer used)"])
    sheet.append(["CHF -> EUR", 0.96])
    sheet.append([])
    sheet.append(["This sheet is kept only for reference. The app now fetches"])
    sheet.append(["a live CHF->EUR rate at generation time -- see fx_rate.py."])


def main() -> None:
    workbook = openpyxl.Workbook()
    build_price_list_sheet(workbook)
    build_legacy_exchange_rate_sheet(workbook)
    workbook.save("sample_data.xlsx")
    total_items = sum(len(items) for _, items in CATEGORIES)
    print(f"Wrote sample_data.xlsx with {len(CATEGORIES)} categories and {total_items} items")


if __name__ == "__main__":
    main()
